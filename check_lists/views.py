from rest_framework import generics
from rest_framework.response import Response

from openpyxl import load_workbook

from .models import CheckList
from .serializers import CheckListNamesSerializer


class CheckListNamesView(generics.ListAPIView):
    serializer_class = CheckListNamesSerializer

    def get_queryset(self):
        type = self.request.query_params.get('type', None)
        return CheckList.objects.filter(type=type).order_by('name')


class CheckListQuestionsView(generics.RetrieveAPIView):
    queryset = CheckList.objects.all()

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        file = instance.file
        name = instance.name

        workbook = load_workbook(filename=file)
        ws = workbook.active

        questions = [cell.value for cell in ws['A'] if cell.value]
        data = {'name': name, 'questions': questions}

        return Response(data)
